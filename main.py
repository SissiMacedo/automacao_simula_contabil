from playwright.sync_api import sync_playwright
import re
from openpyxl import load_workbook
import json

# Lendo os dados do arquivo json
with open('credenciais.json', 'r') as f:
    dados = json.load(f)

    USUARIO = dados['usuario']
    SENHA = dados['senha']

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    context = browser.new_context(
        viewport={"width": 1366, "height": 768},
        accept_downloads=True,
        locale="pt-BR",
        record_video_dir="./video", # Gravação ativa aqui
        record_video_size={"width": 1366, "height": 768} # Opcional: força o tamanho do vídeo
    )
    
    page = context.new_page()
    
    # Abrir o site da Simula Contábil
    page.goto('https://simulacontabil.netlify.app/')

    # Fazer o login com email e senha
    page.get_by_test_id("email-input").click()
    page.get_by_test_id("email-input").type(USUARIO)

    page.get_by_test_id("password-input").click()
    page.get_by_test_id("password-input").type(SENHA)

    page.get_by_test_id("login-button").click()
    # Clicar em Lançamentos
    page.get_by_test_id("nav-lançamentos").click()
    # Carregar a planilha
    planilha = load_workbook('lancamentos.xlsx')
    # Acessar a guia de Lançamentos
    guia_Lancamentos = planilha['Lançamentos']
    # Passar por cada linha da planilha
    for linha in guia_Lancamentos.iter_rows(min_row=2, values_only=True):
        descricao = linha[0]
        valor = str(linha[1])
        data = linha[2]
        tipo = linha[3]
        categoria = linha[4]
        status = linha[5]
        # Clicar em Novo Lançamento
        page.get_by_test_id("btn-new-transaction").click()
        # Preencher Descrição
        page.get_by_test_id("input-description").click()
        page.get_by_test_id("input-description").fill(descricao)
        # Preencher Valor
        page.get_by_test_id("input-amount").click()
        page.get_by_test_id("input-amount").fill(valor)
        # Preencher Data
        page.get_by_test_id("input-date").click()
        page.get_by_test_id("input-date").fill(data)
        # Preencher Tipo
        if tipo == "Receita":
            page.get_by_test_id("select-type").select_option("RECEITA")
        else:
            page.get_by_test_id("select-type").select_option("DESPESA")
        # Preencher Categoria
        page.get_by_test_id("select-category").select_option(categoria)
        # Preencher Status
        page.locator("div").filter(has_text=re.compile(r"^StatusPendentePagoAtrasado$")).get_by_role("combobox").select_option(status.upper())
        # Clicar em Salvar
        page.get_by_test_id("btn-save").click()
        page.wait_for_timeout(2000) # Esperar 2 segundos para garantir que o lançamento seja salvo antes de passar para o próximo
    print("Todos os lançamentos foram inseridos com sucesso!")


    input('Aperte Enter para encerrar a automação...')

    context.close() # O vídeo é processado aqui
    browser.close()